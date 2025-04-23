import pandas as pd
import os
from pathlib import Path
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score, roc_auc_score, accuracy_score

class_for_metrics = 1 # 1 for uncivil and 0 for civil
table_name = "compact_result_table_uncivil_without_duplicates.xlsx" if class_for_metrics == 1 else "compact_result_table_civil_without_duplicates.xlsx"

def get_errors():
    data_path = Path('data')
    duplicates_path = data_path / "all_duplicates.csv"
    duplicates = pd.read_csv(duplicates_path, index_col=0)
    results_path = Path("results")
    models_path = [path for path in results_path.iterdir() if path.is_dir() and path.name not in ["mistral-nemo_12b"]]
    unique_index_error = []

    for model in models_path:
        strategy_path = [path for path in model.iterdir() if path.is_dir()]
        for strategy in strategy_path:
            error_path = strategy / "errors.json"
            error_df = pd.read_json(error_path)
            
            # coletando os indices em que ocorreram erros
            if len(error_df) != 0:
                error_df = error_df.loc[~error_df['index'].isin(duplicates.index.to_list())] # removendo os duplicados
                unique_index_error.extend(list(error_df.loc[:, 'index'].values)) # adicionando os indices dos erros

    return list(set(unique_index_error)) # removendo os indices duplicados

def get_compact_result_table():
    cur_dir = Path(os.getcwd())
    results_dir = cur_dir / 'results'

    # Estratégias únicas
    unique_strategies = [
        "zero_shot", "one_shot", "few_shot", "auto_cot",
        "role_based", "role_based_few_shot",
        "role_based_one_shot", "role_based_auto_cot"
    ]
    
    # Lista de estratégias e modelos
    strategies = []
    models = []
    
    base_models = [
        "deepseek-r1_8b", "deepseek-r1_14b", "gemma_7b", "gemma2_9b",
        "llama3.1_8b", "llama3.2_3b", "mistral_7b", "mistral-nemo_12b",
        "phi4_14b", "gpt-4o-mini"
    ]
    
    # Para cada estratégia, adiciona os modelos
    for strat in unique_strategies:
        strategies.extend([strat] * len(base_models))
        models.extend(base_models)

    # Garantir que os tamanhos batem
    assert len(strategies) == len(models), "strategies e models devem ter o mesmo tamanho"

    # Modelos únicos
    unique_models = np.unique(models)

    # Criando um MultiIndex
    index = pd.MultiIndex.from_tuples(list(zip(strategies, models)), names=["Strategy", "Model"])

    # Criando o DataFrame com o MultiIndex
    result_table = pd.DataFrame(index=index, columns=["precision", "recall", "f1-score", "accuracy", "roc_auc", "FP", "FN"])

    # Preenchendo o DataFrame com os resultados
    for strategy in unique_strategies:
        for model in unique_models:
            model_path = results_dir / model
            strategy_path = model_path / strategy
            predictions_path = strategy_path / f"predictions_df.csv"

            print(predictions_path)
            pred_df = pd.read_csv(predictions_path) 
            #pred_df = pred_df.loc[~pred_df['index'].isin(get_errors())] # removendo os indices dos erros
            pred_df = pred_df.drop_duplicates(subset=['message']) # removendo os duplicados

            # **casos em que há partes da mensagem na predição**
            overwrited = pred_df[~pred_df['prediction'].isin(['0', '1', 0, 1])].copy()
            #print(overwrited)
            if len(overwrited) > 0:
                pred_df = pred_df.drop(overwrited.index, axis=0)
                overwrited['prediction'] = overwrited['actual'].copy()
                overwrited['actual'] = overwrited['source'].copy()
                overwrited['source'] = overwrited['Unnamed: 5']

            pred_df = pd.concat([pred_df, overwrited])
            pred_df['prediction'] = pred_df['prediction'].astype(int)
            pred_df['actual'] = pred_df['actual'].astype(int)

            #print(pred_df.isna().sum())
            f1 = f1_score(pred_df['actual'], pred_df['prediction'], pos_label=class_for_metrics)
            precision = precision_score(pred_df['actual'], pred_df['prediction'], pos_label=class_for_metrics)
            recall = recall_score(pred_df['actual'], pred_df['prediction'], pos_label=class_for_metrics)
            accuracy = accuracy_score(pred_df['actual'], pred_df['prediction'])
            roc_auc = roc_auc_score(pred_df['actual'], pred_df['prediction'])
            fn = confusion_matrix(pred_df['actual'], pred_df['prediction'])[0, 1]
            fp = confusion_matrix(pred_df['actual'], pred_df['prediction'])[1, 0]

            result_table.loc[(strategy, model), "precision"] = np.round(precision, 3)
            result_table.loc[(strategy, model), "recall"] = np.round(recall, 3)
            result_table.loc[(strategy, model), "f1-score"] = np.round(f1, 3)
            result_table.loc[(strategy, model), "accuracy"] = np.round(accuracy, 3)
            result_table.loc[(strategy, model), "roc_auc"] = np.round(roc_auc, 3)

            result_table.loc[(strategy, model), "FP"] = fp 
            result_table.loc[(strategy, model), "FN"] = fn
     
    # Preenchendo os valores do refined model
    refined_model_path = results_dir / "refined_model"
    refined_model_pred_path = refined_model_path / "pred_by_refined_model.csv"
    refined_model_pred_df = pd.read_csv(refined_model_pred_path)
    refined_model_pred_df = refined_model_pred_df.drop_duplicates(subset=['message'])

    f1 = f1_score(refined_model_pred_df['actual'], refined_model_pred_df['pred_by_refined_model'])
    precision = precision_score(refined_model_pred_df['actual'], refined_model_pred_df['pred_by_refined_model'])
    recall = recall_score(refined_model_pred_df['actual'], refined_model_pred_df['pred_by_refined_model'])
    fn = confusion_matrix(refined_model_pred_df['actual'], refined_model_pred_df['pred_by_refined_model'])[0, 1]
    fp = confusion_matrix(refined_model_pred_df['actual'], refined_model_pred_df['pred_by_refined_model'])[1, 0]
    accuracy = accuracy_score(refined_model_pred_df['actual'], refined_model_pred_df['pred_by_refined_model'])
    roc_auc = roc_auc_score(refined_model_pred_df['actual'], refined_model_pred_df['pred_by_refined_model'])

    result_table.loc[("refined_model", "refined_model"), "precision"] = np.round(precision, 3)
    result_table.loc[("refined_model", "refined_model"), "recall"] = np.round(recall, 3)
    result_table.loc[("refined_model", "refined_model"), "f1-score"] = np.round(f1, 3)
    result_table.loc[("refined_model", "refined_model"), "accuracy"] = np.round(accuracy, 3)
    result_table.loc[("refined_model", "refined_model"), "roc_auc"] = np.round(roc_auc, 3)

    result_table.loc[("refined_model", "refined_model"), "FP"] = fp
    result_table.loc[("refined_model", "refined_model"), "FN"] = fn

    # Preenchendo os valores do toxicr
    toxicr_path = results_dir / "toxicr"
    toxicr_pred_path = toxicr_path / "pred_toxicr.csv"
    toxicr_pred_df = pd.read_csv(toxicr_pred_path)
    toxicr_pred_df = toxicr_pred_df.drop_duplicates(subset=['message'])

    f1 = f1_score(toxicr_pred_df['actual'], toxicr_pred_df['pred_by_refined_model'])
    precision = precision_score(toxicr_pred_df['actual'], toxicr_pred_df['pred_by_refined_model'])
    recall = recall_score(toxicr_pred_df['actual'], toxicr_pred_df['pred_by_refined_model'])
    fn = confusion_matrix(toxicr_pred_df['actual'], toxicr_pred_df['pred_by_refined_model'])[0, 1]
    fp = confusion_matrix(toxicr_pred_df['actual'], toxicr_pred_df['pred_by_refined_model'])[1, 0]
    accuracy = accuracy_score(toxicr_pred_df['actual'], toxicr_pred_df['pred_by_refined_model'])
    roc_auc = roc_auc_score(toxicr_pred_df['actual'], toxicr_pred_df['pred_by_refined_model'])

    result_table.loc[("toxicr", "toxicr"), "precision"] = np.round(precision, 3)
    result_table.loc[("toxicr", "toxicr"), "recall"] = np.round(recall, 3)
    result_table.loc[("toxicr", "toxicr"), "f1-score"] = np.round(f1, 3)
    result_table.loc[("toxicr", "toxicr"), "accuracy"] = np.round(accuracy, 3)
    result_table.loc[("toxicr", "toxicr"), "roc_auc"] = np.round(roc_auc, 3)

    result_table.loc[("toxicr", "toxicr"), "FP"] = fp
    result_table.loc[("toxicr", "toxicr"), "FN"] = fn

    
    # Salvar como Excel
    excel_path = results_dir / table_name
    result_table.to_excel(excel_path, index=True)

    # Carregar o arquivo para aplicar estilos
    wb = load_workbook(excel_path)
    ws = wb.active

    # Estilos
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Aplicar negrito nos títulos das colunas
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.border = thin_border  # Aplicar borda

    # Aplicar negrito nos índices e bordas nas células
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).font = bold_font  # Negrito no índice "Strategy"
        ws.cell(row=row, column=2).font = bold_font  # Negrito no índice "Model"
        ws.cell(row=row, column=1).border = thin_border  # Borda no índice
        ws.cell(row=row, column=2).border = thin_border  # Borda no índice

        for col in range(3, ws.max_column + 1):  
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # Adicionar borda
            if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0.7:
                cell.font = bold_font  # Negrito para valores > 0.7

    # Salvar o arquivo Excel com as formatações
    wb.save(excel_path)

get_compact_result_table()
