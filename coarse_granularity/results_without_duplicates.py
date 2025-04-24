import pandas as pd
import os
from pathlib import Path
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from sklearn.metrics import confusion_matrix, f1_score, precision_score, recall_score, roc_auc_score, accuracy_score

class_for_metrics = 1 # 1 for uncivil and 0 for civil
table_name = "compact_result_table_uncivil_without_duplicates.xlsx" if class_for_metrics == 1 else "compact_result_table_civil_without_duplicates.xlsx"

def get_errors():
    data_path = Path('data')
    duplicates_path = data_path / "all_duplicates.csv"
    duplicates = pd.read_csv(duplicates_path, index_col=0)
    results_path = Path("results")
    models_path = [path for path in results_path.iterdir() if path.is_dir() and path.name not in ["mistral-nemo_12b"]]
    unique_index_error = []

    for model in models_path:
        strategy_path = [path for path in model.iterdir() if path.is_dir()]
        for strategy in strategy_path:
            error_path = strategy / "errors.json"
            error_df = pd.read_json(error_path)
            
            # coletando os indices em que ocorreram erros
            if len(error_df) != 0:
                error_df = error_df.loc[~error_df['index'].isin(duplicates.index.to_list())] # removendo os duplicados
                unique_index_error.extend(list(error_df.loc[:, 'index'].values)) # adicionando os indices dos erros

    return list(set(unique_index_error)) # removendo os indices duplicados

def get_compact_result_table():
    cur_dir = Path(os.getcwd())
    results_dir = cur_dir / 'results'

    unique_strategies = [
        "zero_shot", "one_shot", "few_shot", "auto_cot",
        "role_based", "role_based_few_shot",
        "role_based_one_shot", "role_based_auto_cot"
    ]
    
    strategies = []
    models = []
    
    base_models = [
        "deepseek-r1_8b", "deepseek-r1_14b", "gemma_7b", "gemma2_9b",
        "llama3.1_8b", "llama3.2_3b", "mistral_7b", "mistral-nemo_12b",
        "phi4_14b", "gpt-4o-mini"
    ]
    
    # For each strategy, add the models
    for strat in unique_strategies:
        strategies.extend([strat] * len(base_models))
        models.extend(base_models)

    # Ensure sizes match
    assert len(strategies) == len(models), "strategies e models devem ter o mesmo tamanho"

    unique_models = np.unique(models)

    # Building the MultiIndex
    index = pd.MultiIndex.from_tuples(list(zip(strategies, models)), names=["Strategy", "Model"])

    # Building the result table
    result_table = pd.DataFrame(index=index, columns=["precision", "recall", "f1-score", "accuracy", "roc_auc", "FP", "FN"])

    # Filling the result table
    for strategy in unique_strategies:
        for model in unique_models:
            model_path = results_dir / model
            strategy_path = model_path / strategy
            predictions_path = strategy_path / f"predictions_df.csv"

            print(predictions_path)
            pred_df = pd.read_csv(predictions_path) 
            pred_df = pred_df.drop_duplicates(subset=['message']) # removing duplicates

            # **cases where there are parts of the message in the prediction**
            overwrited = pred_df[~pred_df['prediction'].isin(['0', '1', 0, 1])].copy()
            #print(overwrited)
            if len(overwrited) > 0:
                pred_df = pred_df.drop(overwrited.index, axis=0)
                overwrited['prediction'] = overwrited['actual'].copy()
                overwrited['actual'] = overwrited['source'].copy()
                overwrited['source'] = overwrited['Unnamed: 5']

            pred_df = pd.concat([pred_df, overwrited])
            pred_df['prediction'] = pred_df['prediction'].astype(int)
            pred_df['actual'] = pred_df['actual'].astype(int)

            #print(pred_df.isna().sum())
            f1 = f1_score(pred_df['actual'], pred_df['prediction'], pos_label=class_for_metrics)
            precision = precision_score(pred_df['actual'], pred_df['prediction'], pos_label=class_for_metrics)
            recall = recall_score(pred_df['actual'], pred_df['prediction'], pos_label=class_for_metrics)
            accuracy = accuracy_score(pred_df['actual'], pred_df['prediction'])
            roc_auc = roc_auc_score(pred_df['actual'], pred_df['prediction'])
            fn = confusion_matrix(pred_df['actual'], pred_df['prediction'])[0, 1]
            fp = confusion_matrix(pred_df['actual'], pred_df['prediction'])[1, 0]

            result_table.loc[(strategy, model), "precision"] = np.round(precision, 3)
            result_table.loc[(strategy, model), "recall"] = np.round(recall, 3)
            result_table.loc[(strategy, model), "f1-score"] = np.round(f1, 3)
            result_table.loc[(strategy, model), "accuracy"] = np.round(accuracy, 3)
            result_table.loc[(strategy, model), "roc_auc"] = np.round(roc_auc, 3)

            result_table.loc[(strategy, model), "FP"] = fp 
            result_table.loc[(strategy, model), "FN"] = fn
    
    # Save the result table to an Excel file
    excel_path = results_dir / table_name
    result_table.to_excel(excel_path, index=True)

    # load the workbook and select the active worksheet
    wb = load_workbook(excel_path)
    ws = wb.active

    # Styles
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Apply bold font and borders to the header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.border = thin_border  # Aplicar borda

    # Apply bold font and borders to the index columns
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).font = bold_font  
        ws.cell(row=row, column=2).font = bold_font  
        ws.cell(row=row, column=1).border = thin_border  
        ws.cell(row=row, column=2).border = thin_border  

        for col in range(3, ws.max_column + 1):  
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # add border to all cells
            if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0.7:
                cell.font = bold_font  

    # Save excel file with the applied styles
    wb.save(excel_path)

get_compact_result_table()
